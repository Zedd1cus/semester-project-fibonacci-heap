from openpyxl import load_workbook
from benchmark.benchmark import bench_insert, bench_union, bench_extract_min


def get_name_of_txt(number_of_elements: int) -> str:
    return f'/{number_of_elements}.txt'


def get_value(number_of_elements: int, data_number: str) -> int:
    some_dict = get_some_dict(number_of_elements)
    return some_dict[data_number]


def get_some_dict(number: int) -> dict:
    if number == 100:
        return built_dict(2)

    elif number == 5000:
        return built_dict(52)

    elif number == 100000:
        return built_dict(102)

    elif number == 5000000:
        return built_dict(152)


def built_dict(first_number_in_range: int) -> dict:
    some_arr = ['01', '02', '03', '04', '05']
    some_dict = {}
    j = first_number_in_range
    for k in some_arr:
        some_dict[k] = j
        j += 10
    return some_dict


def get_path(number_of_elements: int, data_number: str, path_of_operation) -> str:
    name_of_txt = '/' + str(number_of_elements) + '.txt'
    name_of_dir = '/' + data_number
    path = path_of_operation + name_of_dir + name_of_txt
    return path


def generate_bench_insert(number_of_elements: int, data_number: str) -> None:
    first_number = get_value(number_of_elements, data_number)

    final_path_insert = get_path(number_of_elements, data_number, path_insert)
    final_path_extract_min = get_path(number_of_elements, data_number, path_extract_min)
    final_path_union = get_path(number_of_elements, data_number, path_union)

    count = 0
    while count < 10:
        ws['D' + str(first_number + count)] = bench_insert(final_path_insert)
        ws['E' + str(first_number + count)] = bench_extract_min(final_path_extract_min)
        ws['F' + str(first_number + count)] = bench_union(final_path_union)
        count += 1


path_extract_min = 'C:/ITIS/ASD/1sw/dataset/extractmin'
path_insert = 'C:/ITIS/ASD/1sw/dataset/insert'
path_union = 'C:/ITIS/ASD/1sw/dataset/union'

fn = 'controltests.xlsx'
wb = load_workbook(fn)
ws = wb['first_data']

arr_of_number_of_elements = [100, 5_000, 100_000, 5_000_000]
arr_of_data_numbers = ['01', '02', '03', '04', '05']

for i in arr_of_number_of_elements:
    for j in arr_of_data_numbers:
        generate_bench_insert(i, j)

wb.save(fn)
wb.close()



























