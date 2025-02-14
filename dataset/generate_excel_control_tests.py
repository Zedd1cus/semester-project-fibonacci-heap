from openpyxl import Workbook
from benchmark.benchmark import bench_insert, bench_union, bench_extract_min


def get_value(number_of_elements: int, data_number: str) -> int:
    some_dict = get_some_dict(number_of_elements)
    return some_dict[data_number]


def get_some_dict(number: int) -> dict:
    position = 2
    for num in arr_of_number_of_elements:
        if num == number:
            return built_dict(position)
        position += 50


def built_dict(first_number_in_range: int) -> dict:
    some_arr = ['01', '02', '03', '04', '05']
    some_dict = {}
    j = first_number_in_range
    for k in some_arr:
        some_dict[k] = j
        j += 10
    return some_dict


def get_path(number_of_elements: int, data_number: str, path_of_operation) -> str:
    name_of_txt = '/' + str(number_of_elements) + '.txt'
    name_of_dir = '/' + data_number
    path = path_of_operation + name_of_dir + name_of_txt
    return path


def generate_bench(number_of_elements: int, data_number: str) -> None:
    first_number = get_value(number_of_elements, data_number)

    if data_number == '01':
        ws['A'+str(first_number)] = number_of_elements

    final_path_insert = get_path(number_of_elements, data_number, path_insert)
    final_path_extract_min = get_path(number_of_elements, data_number, path_extract_min)
    final_path_union = get_path(number_of_elements, data_number, path_union)

    ws['B' + str(first_number)] = data_number

    count = 0
    while count < 10:
        ws['D' + str(first_number + count)] = bench_insert(final_path_insert)
        ws['E' + str(first_number + count)] = bench_extract_min(final_path_extract_min)
        ws['F' + str(first_number + count)] = bench_union(final_path_union)
        count += 1


if __name__ == '__main__':
    ex_file = Workbook()
    ws = ex_file.create_sheet(title='first_data', index=0)

    path_extract_min = 'extractmin'
    path_insert = 'insert'
    path_union = 'union'

    arr_of_number_of_elements = [100, 500, 900, 1300, 1700, 2100, 2500, 2900]
    arr_of_data_numbers = ['01', '02', '03', '04', '05']

    ws['A1'] = 'КОЛИЧЕСТВО ЭЛЕМЕНТОВ'
    ws['B1'] = 'НАБОР ДАННЫХ'
    ws['C1'] = 'НОМЕР ЗАПУСКА'
    ws['D1'] = 'INSERT'
    ws['E1'] = 'EXTRACT_MIN'
    ws['F1'] = 'UNION'
    ws['I1'] = 'Среднее время выполнения'
    ws['J1'] = 'В МС'
    ws['I2'] = 'Размер набора данных'
    ws['J2'] = 'INSERT'
    ws['K2'] = 'EXTRACT_MIN'
    ws['L2'] = 'UNION'

    for i in arr_of_number_of_elements:
        for j in arr_of_data_numbers:
            generate_bench(i, j)

    for i in range(len(arr_of_number_of_elements)):
        ws['I' + str(3+i)] = arr_of_number_of_elements[i]

        frst_position = get_value(arr_of_number_of_elements[i], '01')

        ws['J' + str(3+i)] = f'=СРЗНАЧ(D{frst_position}:D{frst_position+49})'
        ws['K' + str(3+i)] = f'=СРЗНАЧ(E{frst_position}:E{frst_position+49})'
        ws['L' + str(3+i)] = f'=СРЗНАЧ(F{frst_position}:F{frst_position+49})'

    ex_file.save(filename='controltests.xlsx')
    ex_file.close()